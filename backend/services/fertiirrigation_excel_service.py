"""
FertiIrrigation Excel Export Service.
Generates professional Excel reports for fertigation calculations.
"""
from io import BytesIO
from datetime import datetime
from typing import Dict, Any, Optional, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from app.services.fertiirrigation_ab_tanks_service import separate_fertilizers_ab

FERTIRRIEGO_GREEN = "10B981"
FERTIRRIEGO_DARK = "059669"
HEADER_BG = "D1FAE5"


class FertiIrrigationExcelService:
    """Service for generating FertiIrrigation Excel reports."""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color=FERTIRRIEGO_DARK, end_color=FERTIRRIEGO_DARK, fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.title_font = Font(bold=True, size=16, color=FERTIRRIEGO_DARK)
        self.subtitle_font = Font(bold=True, size=12, color=FERTIRRIEGO_DARK)
        self.light_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _apply_header_style(self, ws, row_num: int, max_col: int):
        """Apply header styling to a row."""
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
    
    def _apply_border_to_range(self, ws, start_row: int, start_col: int, end_row: int, end_col: int):
        """Apply borders to a range of cells."""
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = self.border
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 12), 40)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate_fertiirrigation_excel(
        self,
        calculation: Dict[str, Any],
        user_name: str = "Usuario",
        extraction_curve_info: Optional[Dict] = None
    ) -> BytesIO:
        """
        Generate Excel report for a fertigation calculation.
        
        Args:
            calculation: The calculation data with results
            user_name: Name of the user
            extraction_curve_info: Optional extraction curve data
        
        Returns:
            BytesIO with Excel file content
        """
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        
        result = calculation.get('result', calculation.get('results', {}))
        
        ws_summary = self._create_summary_sheet(wb, calculation, user_name, extraction_curve_info)
        ws_balance = self._create_nutrient_balance_sheet(wb, result)
        ws_program = self._create_fertilizer_program_sheet(wb, result, calculation)
        self._create_acid_sheet(wb, result, calculation)
        self._create_soil_depletion_sheet(wb, result)
        self._create_ab_tanks_sheet(wb, result, calculation)
        if extraction_curve_info:
            self._create_extraction_curve_sheet(wb, extraction_curve_info)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _create_summary_sheet(self, wb, calculation: Dict, user_name: str, extraction_info: Optional[Dict]) -> Any:
        """Create the summary sheet."""
        ws = wb.create_sheet("Resumen")
        row = 1
        
        ws.cell(row=row, column=1, value="REPORTE DE FERTIRRIEGO - AGRIDOSER").font = self.title_font
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        ws.cell(row=row, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(italic=True)
        row += 2
        
        ws.cell(row=row, column=1, value="INFORMACIÓN GENERAL").font = self.subtitle_font
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        irrigation_frequency_days = calculation.get(
            'irrigation_frequency_days',
            calculation.get('irrigation', {}).get('irrigation_frequency_days')
        )
        irrigation_volume_m3_ha = calculation.get(
            'irrigation_volume_m3_ha',
            calculation.get('irrigation', {}).get('irrigation_volume_m3_ha')
        )
        stage_duration = extraction_info.get('duration_days') if extraction_info else None

        info_data = [
            ("Nombre del Cálculo:", calculation.get('name', 'Sin nombre')),
            ("Usuario:", user_name),
            ("Cultivo:", calculation.get('crop_name', 'N/A')),
            ("Área (ha):", calculation.get('area_ha', 1.0)),
            ("Análisis de Suelo:", calculation.get('soil_analysis_name', 'N/A')),
            ("Análisis de Agua:", calculation.get('water_analysis_name', 'N/A')),
            ("N° Aplicaciones:", calculation.get('num_applications', 10)),
        ]

        if irrigation_frequency_days:
            info_data.append(("Frecuencia de riego (días):", irrigation_frequency_days))
        if irrigation_volume_m3_ha:
            info_data.append(("Volumen de riego (m³/ha):", irrigation_volume_m3_ha))
        if stage_duration:
            if isinstance(stage_duration, dict):
                min_days = stage_duration.get('min')
                max_days = stage_duration.get('max')
                if min_days is not None and max_days is not None:
                    duration_label = f"{min_days}-{max_days}"
                else:
                    duration_label = str(stage_duration)
            else:
                duration_label = str(stage_duration)
            info_data.append(("Duración de etapa (días):", duration_label))
        
        if extraction_info:
            info_data.append(("Curva de Extracción:", f"{extraction_info.get('crop_name', '')} - {extraction_info.get('stage_name', '')}"))
        
        for label, value in info_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = self.light_fill
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=1).border = self.border
            ws.cell(row=row, column=2).border = self.border
            row += 1
        
        row += 1
        
        ws.cell(row=row, column=1, value="RESUMEN NUTRICIONAL (kg/ha)").font = self.subtitle_font
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        result = calculation.get('result', calculation.get('results', {}))
        nutrient_summary = [
            ("Nitrógeno (N):", result.get('total_n_kg_ha', 0)),
            ("Fósforo (P₂O₅):", result.get('total_p2o5_kg_ha', 0)),
            ("Potasio (K₂O):", result.get('total_k2o_kg_ha', 0)),
            ("Calcio (Ca):", result.get('total_ca_kg_ha', 0)),
            ("Magnesio (Mg):", result.get('total_mg_kg_ha', 0)),
            ("Azufre (S):", result.get('total_s_kg_ha', 0)),
        ]
        
        for label, value in nutrient_summary:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = self.light_fill
            ws.cell(row=row, column=2, value=round(value, 2) if value else 0)
            ws.cell(row=row, column=1).border = self.border
            ws.cell(row=row, column=2).border = self.border
            row += 1
        
        self._auto_adjust_columns(ws)
        return ws
    
    def _create_nutrient_balance_sheet(self, wb, result: Dict) -> Any:
        """Create the nutrient balance sheet."""
        ws = wb.create_sheet("Balance Nutricional")
        nutrient_balance = result.get('nutrient_balance', [])
        has_acid = any((nb.get('acid_contribution_kg_ha', 0) or 0) > 0 for nb in nutrient_balance)
        
        headers = ["Nutriente", "Requerimiento (kg/ha)", "Disponible Suelo", "Aporte Agua"]
        if has_acid:
            headers.append("Aporte Ácido")
        headers += ["Déficit", "Eficiencia (%)", "A Aplicar (kg/ha)"]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))
        
        row = 2
        for nb in nutrient_balance:
            ws.cell(row=row, column=1, value=nb.get('nutrient', ''))
            ws.cell(row=row, column=2, value=round(nb.get('requirement_kg_ha', 0), 2))
            ws.cell(
                row=row,
                column=3,
                value=round(
                    nb.get('soil_contribution_kg_ha',
                           nb.get('soil_diagnostic_kg_ha', nb.get('soil_available_kg_ha', 0))),
                    2
                )
            )
            ws.cell(row=row, column=4, value=round(nb.get('water_contribution_kg_ha', 0), 2))
            col_offset = 0
            if has_acid:
                ws.cell(row=row, column=5, value=round(nb.get('acid_contribution_kg_ha', 0), 2))
                col_offset = 1
            ws.cell(row=row, column=5 + col_offset, value=round(nb.get('deficit_kg_ha', 0), 2))
            ws.cell(row=row, column=6 + col_offset, value=round(nb.get('efficiency_factor', 1) * 100, 0))
            ws.cell(row=row, column=7 + col_offset, value=round(nb.get('fertilizer_needed_kg_ha', 0), 2))
            
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = self.border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            row += 1
        
        if len(nutrient_balance) > 0:
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "clustered"
            chart.style = 10
            chart.title = "Balance Nutricional Completo"
            chart.y_axis.title = "kg/ha"
            chart.x_axis.title = "Nutriente"
            
            data = Reference(ws, min_col=2, max_col=5, min_row=1, max_row=row-1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=row-1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            chart.width = 18
            chart.height = 12
            
            ws.add_chart(chart, "I2")
        
        self._auto_adjust_columns(ws)
        return ws

    def _create_acid_sheet(self, wb, result: Dict, calculation: Dict) -> Optional[Any]:
        """Create the acid treatment sheet when acid is present."""
        acid_treatment = result.get('acid_treatment')
        acid_program = result.get('acid_program')

        if not acid_treatment and acid_program:
            ml_per_1000L = acid_program.get('ml_per_1000L', 0)
            acid_treatment = {
                'acid_id': acid_program.get('acid_id', ''),
                'acid_name': acid_program.get('acid_name', ''),
                'acid_type': acid_program.get('acid_id', ''),
                'ml_per_1000L': ml_per_1000L,
                'cost_per_1000L': acid_program.get('cost_per_1000L', 0),
                'nutrient_contribution': acid_program.get('nutrient_contribution', {}),
            }

        if not acid_treatment:
            return None

        if not (acid_treatment.get('acid_name') or acid_treatment.get('ml_per_1000L', 0) > 0 or acid_treatment.get('volume_l_ha', 0) > 0):
            return None

        ws = wb.create_sheet("Tratamiento Ácido")

        num_applications = calculation.get('num_applications', result.get('num_applications', 10))
        area_ha = calculation.get('area_ha', result.get('area_ha', 1.0))
        irrigation_volume_m3_ha = result.get('irrigation_volume_m3_ha', calculation.get('irrigation_volume_m3_ha', 0))
        water_volume_1000L = calculation.get('water_volume_1000L', result.get('water_volume_1000L', 0))
        total_water_1000L = water_volume_1000L or (irrigation_volume_m3_ha * num_applications)

        nutrient_contrib = acid_treatment.get('nutrient_contribution', {})
        nutrient_in_kg_ha = acid_treatment.get('nutrient_in_kg_ha', False)

        if nutrient_in_kg_ha:
            acid_n = nutrient_contrib.get('N', nutrient_contrib.get('NO3_N', 0)) * area_ha
            acid_p = nutrient_contrib.get('P', nutrient_contrib.get('P2O5', 0)) * area_ha
            acid_s = nutrient_contrib.get('S', nutrient_contrib.get('SO4_S', 0)) * area_ha
        elif nutrient_contrib:
            acid_n = (nutrient_contrib.get('N', nutrient_contrib.get('NO3_N', 0)) * total_water_1000L) / 1000.0
            acid_p = (nutrient_contrib.get('P', nutrient_contrib.get('P2O5', 0)) * total_water_1000L) / 1000.0
            acid_s = (nutrient_contrib.get('S', nutrient_contrib.get('SO4_S', 0)) * total_water_1000L) / 1000.0
        else:
            acid_n = acid_treatment.get('n_kg_ha', 0) * area_ha
            acid_p = acid_treatment.get('p2o5_kg_ha', 0) * area_ha
            acid_s = acid_treatment.get('s_kg_ha', 0) * area_ha

        acid_name = acid_treatment.get('acid_name', 'Ácido')
        acid_id = acid_treatment.get('acid_id', acid_treatment.get('acid_type', ''))
        acid_vol_per_ha = acid_treatment.get('volume_l_ha', 0)
        acid_vol_total = acid_vol_per_ha * area_ha
        acid_vol_per_app = acid_vol_total / num_applications if num_applications > 0 else acid_vol_total
        ml_per_1000L = acid_treatment.get('ml_per_1000L', 0)

        row = 1
        ws.cell(row=row, column=1, value="TRATAMIENTO CON ÁCIDO").font = self.title_font
        ws.merge_cells(f'A{row}:B{row}')
        row += 2

        data = [
            ("Ácido utilizado:", acid_name),
            ("ID de ácido:", acid_id),
        ]
        if ml_per_1000L:
            data.append(("Dosis de inyección:", f"{ml_per_1000L:.1f} ml/1000L"))
        data.extend([
            ("Dosis por hectárea:", f"{acid_vol_per_ha:.2f} L/ha"),
            ("Dosis total:", f"{acid_vol_total:.2f} L"),
            ("Dosis por riego:", f"{acid_vol_per_app:.3f} L"),
        ])

        if acid_n > 0:
            data.append(("Aporte de N:", f"{acid_n:.2f} kg"))
        if acid_p > 0:
            data.append(("Aporte de P₂O₅:", f"{acid_p:.2f} kg"))
        if acid_s > 0:
            data.append(("Aporte de S:", f"{acid_s:.2f} kg"))

        for label, value in data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = self.light_fill
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=1).border = self.border
            ws.cell(row=row, column=2).border = self.border
            row += 1

        self._auto_adjust_columns(ws)
        return ws

    def _create_soil_depletion_sheet(self, wb, result: Dict) -> Optional[Any]:
        """Create the soil depletion sheet when data is available."""
        nutrient_balance = result.get('nutrient_balance', [])
        has_soil_depletion = any(
            nb.get('soil_total_kg_ha', 0) > 0 or nb.get('soil_remaining_kg_ha', 0) > 0
            for nb in nutrient_balance
        )
        if not has_soil_depletion:
            return None

        ws = wb.create_sheet("Agotamiento Suelo")
        headers = ["Nutriente", "Total Disp. (kg/ha)", "Consumido Antes (kg/ha)", "Consumido Etapa (kg/ha)", "Restante (kg/ha)", "% Acum."]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))

        row = 2
        for nb in nutrient_balance:
            ws.cell(row=row, column=1, value=nb.get('nutrient', ''))
            ws.cell(row=row, column=2, value=round(nb.get('soil_total_kg_ha', 0), 2))
            ws.cell(row=row, column=3, value=round(nb.get('soil_consumed_before_kg_ha', 0), 2))
            ws.cell(row=row, column=4, value=round(nb.get('soil_consumed_this_stage_kg_ha', 0), 2))
            ws.cell(row=row, column=5, value=round(nb.get('soil_remaining_kg_ha', 0), 2))
            ws.cell(row=row, column=6, value=f"{nb.get('cumulative_extraction_pct', 0):.1f}%")

            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = self.border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            row += 1

        self._auto_adjust_columns(ws)
        return ws
    
    def _create_fertilizer_program_sheet(self, wb, result: Dict, calculation: Dict) -> Any:
        """Create the fertilizer program sheet."""
        ws = wb.create_sheet("Programa Fertilización")
        
        num_applications = calculation.get('num_applications', result.get('num_applications', 10))
        currency = calculation.get('currency', result.get('currency', 'MXN'))
        currency_symbols = {'MXN': '$', 'USD': '$', 'PEN': 'S/'}
        symbol = currency_symbols.get(currency, '$')
        
        ws.cell(row=1, column=1, value=f"PROGRAMA DE FERTILIZACIÓN ({num_applications} Aplicaciones)").font = self.subtitle_font
        ws.merge_cells('A1:D1')
        
        headers = ["Fertilizante", "Dosis Total (kg/ha)", "Por Riego (kg)", f"Costo ({currency})"]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=header)
        self._apply_header_style(ws, 2, len(headers))
        
        fertilizer_program = result.get('fertilizer_program', [])
        if not fertilizer_program:
            fertilizer_program = calculation.get('fertilizer_program', [])
        
        consolidated_ferts = {}
        for fd in fertilizer_program:
            name = fd.get('fertilizer_name', '')
            if name in consolidated_ferts:
                consolidated_ferts[name]['dose_kg_ha'] += fd.get('dose_kg_ha', 0)
                consolidated_ferts[name]['cost_total'] += fd.get('cost_total', fd.get('cost_ha', 0))
            else:
                consolidated_ferts[name] = {
                    'fertilizer_name': name,
                    'dose_kg_ha': fd.get('dose_kg_ha', 0),
                    'cost_total': fd.get('cost_total', fd.get('cost_ha', 0))
                }
        
        unique_fertilizers = list(consolidated_ferts.values())
        
        row = 3
        total_dose = 0
        total_cost = 0
        
        for fd in unique_fertilizers:
            dose = fd.get('dose_kg_ha', 0)
            dose_per_app = dose / num_applications if num_applications > 0 else dose
            cost = fd.get('cost_total', 0)
            total_dose += dose
            total_cost += cost
            
            ws.cell(row=row, column=1, value=fd.get('fertilizer_name', ''))
            ws.cell(row=row, column=2, value=round(dose, 1))
            ws.cell(row=row, column=3, value=round(dose_per_app, 2))
            ws.cell(row=row, column=4, value=f"{symbol}{cost:.2f}")
            
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = self.border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            row += 1
        
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{total_dose:.1f} kg").font = Font(bold=True)
        ws.cell(row=row, column=3, value=f"{total_dose / num_applications:.2f} kg").font = Font(bold=True)
        ws.cell(row=row, column=4, value=f"{symbol}{total_cost:.2f}").font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = self.border
            ws.cell(row=row, column=col).fill = self.light_fill
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        
        self._auto_adjust_columns(ws)
        return ws
    
    def _create_extraction_curve_sheet(self, wb, extraction_info: Dict) -> Any:
        """Create the extraction curve sheet."""
        ws = wb.create_sheet("Curva Extracción")
        
        row = 1
        ws.cell(row=row, column=1, value="CURVA DE EXTRACCIÓN APLICADA").font = self.subtitle_font
        ws.merge_cells(f'A{row}:C{row}')
        row += 2
        
        ws.cell(row=row, column=1, value="Cultivo:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=extraction_info.get('crop_name', 'N/A'))
        row += 1
        
        ws.cell(row=row, column=1, value="Etapa Fenológica:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=extraction_info.get('stage_name', 'N/A'))
        row += 2
        
        headers = ["Nutriente", "% Absorción Acumulada"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, len(headers))
        row += 1
        
        percentages = extraction_info.get('percentages', {})
        nutrient_display = [
            ("N", "Nitrógeno (N)"),
            ("P2O5", "Fósforo (P₂O₅)"),
            ("K2O", "Potasio (K₂O)"),
            ("Ca", "Calcio (Ca)"),
            ("Mg", "Magnesio (Mg)"),
            ("S", "Azufre (S)")
        ]
        
        norm_pct = {k.upper(): v for k, v in percentages.items()}
        for key, name in nutrient_display:
            pct = norm_pct.get(key.upper(), percentages.get(key, 0))
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=f"{pct}%" if pct > 0 else "N/A")
            ws.cell(row=row, column=1).border = self.border
            ws.cell(row=row, column=2).border = self.border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            row += 1
        
        self._auto_adjust_columns(ws)
        return ws
    
    def _create_ab_tanks_sheet(self, wb, result: Dict, calculation: Dict) -> Any:
        """Create the A/B tanks sheet with fertilizer separation."""
        ws = wb.create_sheet("Tanques A-B")
        
        num_applications = calculation.get('num_applications', result.get('num_applications', 10))
        
        fertilizer_program = result.get('fertilizer_program', [])
        if not fertilizer_program:
            fertilizer_program = calculation.get('fertilizer_program', [])
        
        # Use pre-computed A/B tank data if available (from 'tanks' or 'ab_tanks_separation')
        ab_tanks_precomputed = result.get('tanks') or result.get('ab_tanks_separation')
        if ab_tanks_precomputed:
            tank_a = ab_tanks_precomputed.get('tank_a', [])
            tank_b = ab_tanks_precomputed.get('tank_b', [])
        else:
            # Compute from fertilizer program - pass full dicts to preserve all metadata
            acid_treatment = result.get('acid_treatment')
            ab_tanks_data = separate_fertilizers_ab(fertilizer_program, acid_treatment)
            tank_a = ab_tanks_data.get('tank_a', [])
            tank_b = ab_tanks_data.get('tank_b', [])
        
        row = 1
        ws.cell(row=row, column=1, value="PROGRAMA DE INYECCIÓN - TANQUES A/B").font = self.title_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ws.cell(row=row, column=1, value="Separación de fertilizantes por compatibilidad química").font = Font(italic=True)
        row += 2
        
        tank_a_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        tank_a_light = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        
        ws.cell(row=row, column=1, value="TANQUE A - Calcio + Micronutrientes").font = self.subtitle_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        headers_a = ["Fertilizante", "Dosis Total (kg/ha)", "Por Riego (kg)", "Observación"]
        for col, header in enumerate(headers_a, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = tank_a_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        row += 1
        
        if tank_a:
            for fert in tank_a:
                dose = fert.get('dose_kg_ha', 0)
                dose_per_app = dose / num_applications if num_applications > 0 else dose
                ws.cell(row=row, column=1, value=fert.get('fertilizer_name', ''))
                ws.cell(row=row, column=2, value=round(dose, 2))
                ws.cell(row=row, column=3, value=round(dose_per_app, 3))
                ws.cell(row=row, column=4, value="Calcio / Micro")
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = self.border
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
                    if row % 2 == 0:
                        ws.cell(row=row, column=col).fill = tank_a_light
                row += 1
        else:
            ws.cell(row=row, column=1, value="Sin fertilizantes en Tanque A")
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
        
        row += 1
        
        tank_b_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        tank_b_light = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
        
        ws.cell(row=row, column=1, value="TANQUE B - Fosfatos + Sulfatos + Ácidos").font = self.subtitle_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        headers_b = ["Fertilizante", "Dosis Total", "Por Riego", "Observación"]
        for col, header in enumerate(headers_b, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = tank_b_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        row += 1
        
        if tank_b:
            for fert in tank_b:
                dose = fert.get('dose_kg_ha', 0)
                dose_per_app = dose / num_applications if num_applications > 0 else dose
                unit = "kg/ha" if not fert.get('is_acid') else "L/ha"
                unit_per_app = "kg" if not fert.get('is_acid') else "L"
                obs = "Ácido" if fert.get('is_acid') else "Fosfato/Sulfato"
                ws.cell(row=row, column=1, value=fert.get('fertilizer_name', ''))
                ws.cell(row=row, column=2, value=f"{dose:.2f} {unit}")
                ws.cell(row=row, column=3, value=f"{dose_per_app:.3f} {unit_per_app}")
                ws.cell(row=row, column=4, value=obs)
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = self.border
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
                    if row % 2 == 0:
                        ws.cell(row=row, column=col).fill = tank_b_light
                row += 1
        else:
            ws.cell(row=row, column=1, value="Sin fertilizantes en Tanque B")
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
        
        row += 2
        ws.cell(row=row, column=1, value="⚠ IMPORTANTE: No mezclar fertilizantes de Tanque A con Tanque B.").font = Font(bold=True, color="B45309")
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws.cell(row=row, column=1, value="El calcio precipita con fosfatos y sulfatos, causando obstrucciones.").font = Font(italic=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        self._auto_adjust_columns(ws)
        return ws


fertiirrigation_excel_service = FertiIrrigationExcelService()
